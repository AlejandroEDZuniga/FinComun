import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
import re
import shutil
import dateparser
import datetime

current_directory = os.getcwd()
ruta_imagen = os.path.join(current_directory, 'FC_Logo.png')


def insertar_logo_en_reportes(primera_hoja):
    imagen = Image(ruta_imagen)
    imagen.width = imagen.width * 0.35
    imagen.height = imagen.height * 0.25
    primera_hoja.add_image(imagen, 'A1')

def limpiar_titulo(titulo):
    titulo_limpio = re.sub(r'[<>:"/\\|?*]', '_', titulo)
    return titulo_limpio

def parsear_fecha(fecha):
    meses_espanol = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4, 'Mayo': 5, 'Junio': 6,
        'Julio': 7, 'Agosto': 8, 'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }

    partes_fecha = fecha.split(' ')

    dia = int(partes_fecha[0])
    mes_espanol = partes_fecha[1].capitalize()  # Convertir la primera letra en mayúscula
    anio = int(partes_fecha[2])

    # Obtener el número del mes en español
    mes_numero = meses_espanol[mes_espanol]

    fecha_formateada = datetime.date(anio, mes_numero, dia)
    return fecha_formateada


def unificar_guardar_archivos_RCOBO(ruta_carpeta_acumulado='', archivos_RCOBO=[]):
    libro_excel = Workbook()

    for i, (archivo_RCOBO, ruta_archivo_RCOBO) in enumerate(archivos_RCOBO):
        datos_archivo_RCOBO = pd.read_excel(ruta_archivo_RCOBO, header=None)
        fecha_operacion = datos_archivo_RCOBO.iloc[2, 1]  # Obtener el valor de la celda B3
        fecha_operacion_limpia = fecha_operacion.strftime('%d/%m/%Y')
        datos_archivo_RCOBO.at[2, 1] = fecha_operacion_limpia
        hoja = libro_excel.create_sheet(title=limpiar_titulo(fecha_operacion_limpia), index=i)

        for _, row in datos_archivo_RCOBO.iterrows():
            hoja.append(row.tolist())

        # Ajustar ancho de las columnas en la hoja actual
        for column in hoja.columns:
            column_letter = column[0].column_letter
            hoja.column_dimensions[column_letter].width = 25

        hoja.row_dimensions[1].height = 35

        # Escribir y formatear el título en la primera celda
        primera_celda = hoja.cell(row=1, column=3)
        primera_celda.value = "Reporte de la Cuenta Operativa Banxico Acumulado"
        primera_celda.font = Font(bold=True, size=20)

        alineacion_titulo = Alignment(vertical="center")
        primera_celda.alignment = alineacion_titulo

        insertar_logo_en_reportes(hoja)

    del libro_excel['Sheet']

    # Guardar el archivo Excel unificado
    ruta_archivo_unificado = os.path.join(ruta_carpeta_acumulado, 'RCOBA.xlsx')
    libro_excel.save(ruta_archivo_unificado)
    #print(f"Se ha guardado el archivo unificado 'RCOBA.xlsx' en la ruta: {ruta_archivo_unificado}")

def unificar_guardar_archivos_RTERO(ruta_carpeta_acumulado='', archivos_RTERO=[]):
    libro_excel = Workbook()
    
    for i, (archivo_RTERO, ruta_archivo_RTERO) in enumerate(archivos_RTERO):
        datos_archivo_RTERO = pd.read_excel(ruta_archivo_RTERO, header=None)
        fecha_operacion = datos_archivo_RTERO.iloc[2, 1]  # Obtener el valor de la celda B3
        fecha_operacion_limpia = fecha_operacion.strftime('%d/%m/%Y')
        datos_archivo_RTERO.at[2, 1] = fecha_operacion_limpia
        hoja = libro_excel.create_sheet(title=limpiar_titulo(fecha_operacion_limpia), index=i)

        for _, row in datos_archivo_RTERO.iterrows():
            hoja.append(row.tolist())

        # Ajustar ancho de las columnas en la hoja actual
        for column in hoja.columns:
            column_letter = column[0].column_letter
            hoja.column_dimensions[column_letter].width = 25

        hoja.row_dimensions[1].height = 35

        if hoja['D1'].value == "Reporte de transacciones Enviadas y Recibidas por Operación":
            hoja['D1'].value = None

        # Escribir y formatear el título en la primera celda
        primera_celda = hoja.cell(row=1, column=3)
        primera_celda.value = "Reporte de Transacciones Enviadas y Recibidas Acumulado"
        primera_celda.font = Font(bold=True, size=20)

        alineacion_titulo = Alignment(vertical="center")
        primera_celda.alignment = alineacion_titulo

        insertar_logo_en_reportes(hoja)

    del libro_excel['Sheet']

    ruta_archivo_unificado = os.path.join(ruta_carpeta_acumulado, 'RTERA.xlsx')
    libro_excel.save(ruta_archivo_unificado)
    #print(f"Se ha guardado el archivo unificado 'RTERA.xlsx' en la ruta: {ruta_archivo_unificado}")


def unificar_guardar_archivos_RDSERO(ruta_carpeta_acumulado='', archivos_RDSERO=[]):
    libro_excel = Workbook()
    
    for i, (archivo_RDSERO, ruta_archivo_RDSERO) in enumerate(archivos_RDSERO):
        datos_archivo_RDSERO = pd.read_excel(ruta_archivo_RDSERO, header=None)
        fecha_operacion = datos_archivo_RDSERO.iloc[2, 1]  # Obtener el valor de la celda B3
        fecha_operacion_limpia = fecha_operacion.strftime('%d/%m/%Y')
        datos_archivo_RDSERO.at[2, 1] = fecha_operacion_limpia       
        hoja = libro_excel.create_sheet(title=limpiar_titulo(fecha_operacion_limpia), index=i)

        for _, row in datos_archivo_RDSERO.iterrows():
            hoja.append(row.tolist())

        for column in hoja.columns:
            column_letter = column[0].column_letter
            hoja.column_dimensions[column_letter].width = 25

        hoja.row_dimensions[1].height = 35

        # Escribir y formtear el titulo
        primera_celda = hoja.cell(row=1, column=3)
        primera_celda.value = "Reporte de Diferencia SPEI's Enviados y Recibidos Acumulado"
        primera_celda.font = Font(bold=True, size=20)

        alineacion_titulo = Alignment(vertical="center")
        primera_celda.alignment = alineacion_titulo
    
        insertar_logo_en_reportes(hoja)

    del libro_excel['Sheet']

    ruta_archivo_unificado = os.path.join(ruta_carpeta_acumulado, 'RDSERA.xlsx')
    libro_excel.save(ruta_archivo_unificado)
    #print(f"Se ha guardado el archivo unificado 'RDSERA.xlsx' en la ruta: {ruta_archivo_unificado}")


def unificar_guardar_archivos_RDSERO2(ruta_carpeta_acumulado='', archivos_RDSERO2=[]):
    libro_excel = Workbook()
    
    for i, (archivos_RDSERO2, ruta_archivos_RDSERO2) in enumerate(archivos_RDSERO2):
        datos_archivos_RDSERO2 = pd.read_excel(ruta_archivos_RDSERO2, header=None)
        fecha_operacion = datos_archivos_RDSERO2.iloc[2, 1]  # Obtener el valor de la celda B3
        fecha_operacion_limpia = fecha_operacion.strftime('%d/%m/%Y')
        datos_archivos_RDSERO2.at[2, 1] = fecha_operacion_limpia
        hoja = libro_excel.create_sheet(title=limpiar_titulo(fecha_operacion_limpia), index=i)


        for _, row in datos_archivos_RDSERO2.iterrows():
            hoja.append(row.tolist())

        for column in hoja.columns:
            column_letter = column[0].column_letter
            hoja.column_dimensions[column_letter].width = 25

        hoja.row_dimensions[1].height = 35


        # Escribir y formtear el titulo
        primera_celda = hoja.cell(row=1, column=3)
        primera_celda.value = "Reporte de Diferencia SPEI's Enviados y Recibidos Acumulado 2"
        primera_celda.font = Font(bold=True, size=20)

        alineacion_titulo = Alignment(vertical="center")
        primera_celda.alignment = alineacion_titulo

        insertar_logo_en_reportes(hoja)

    del libro_excel['Sheet']

    ruta_archivo_unificado = os.path.join(ruta_carpeta_acumulado, 'RDSERA2.xlsx')
    libro_excel.save(ruta_archivo_unificado)
    #print(f"Se ha guardado el archivo unificado 'RDSERA2.xlsx' en la ruta: {ruta_archivo_unificado}")


def obtener_archivos_sistemas_flex_karpay_para_procesar_acumulados(fecha_inicio, fecha_fin):
    try:
        ruta_archivos_acumulados_originales = r'C:\Users\svcmerr\Desktop\Acumulados-original'
        ruta_destino_acumulados = r'C:\Users\svcmerr\Desktop\Acumulados'

        fecha_inicio = dateparser.parse(fecha_inicio)
        fecha_fin = dateparser.parse(fecha_fin)

        carpetas_origen = os.listdir(ruta_archivos_acumulados_originales)
        carpetas_destino = os.listdir(ruta_destino_acumulados)

        for carpeta_origen in carpetas_origen:
            carpeta_destino = carpeta_origen.lower()
            if carpeta_destino in carpetas_destino:
                ruta_origen_carpeta = os.path.join(ruta_archivos_acumulados_originales, carpeta_origen)
                ruta_destino_carpeta = os.path.join(ruta_destino_acumulados, carpeta_destino)

                carpetas_reportes = ['REPORTES FLEX', 'REPORTES KARPAY']
                for carpeta_reporte in carpetas_reportes:
                    ruta_origen_reporte = os.path.join(ruta_origen_carpeta, carpeta_reporte)
                    ruta_destino_reporte = os.path.join(ruta_destino_carpeta, carpeta_reporte)

                    if os.path.exists(ruta_origen_reporte):
                        shutil.copytree(ruta_origen_reporte, ruta_destino_reporte, dirs_exist_ok=True)
                        #print(f"Copiando carpeta: {ruta_origen_reporte} -> {ruta_destino_reporte}")
                    else:
                        pass
                        #print(f"No se encontró la carpeta: {ruta_origen_reporte}")

        
    except Exception as e:
        pass
        #print("Error en obtener_archivos_sistemas_flex_karpay_para_procesar_acumulados:", str(e))


#obtener_archivos_sistemas_flex_karpay_para_procesar_acumulados(fecha_inicio='07 Febrero 2023', fecha_fin='09 Febrero 2023')